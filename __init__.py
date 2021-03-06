# Imports
# Necessary to work. If not installed, try " pip3 install pandas"
import pandas as pd
# Necessary to work. If not installed, try " pip3 install xlrd"
import xlrd
from UtilsFunctions import pt
# To delete nan values and work with scientist library
import numpy as np
# To write to an existing file. If not installed, try " pip3 install openpyxl"
import openpyxl
# OS
import os
# TO errors
import traceback


# Paths
path_origin = "D:\\Machine_Learning\\DataSets\\VF\\CARMONA_1_SEMANA_OCTUBRE_(200_REGISTROS).xlsx"
path_technology = "D:\\Machine_Learning\\DataSets\\VF\\neba.xlsx"
paths = [path_origin, path_technology]

#Columns
origin_column_municipality = "POBLACION"
technology_column_municipality = "MUNICIPIO"
origin_column_address = "DIRECCION"
technology_column_address = "Direccion"


origin_sheet = "Hoja1"
technology_sheet = "neba"


def algorithm(paths):
    """
    Execute all algorithm's steps.
    :param paths: 
    :return: 
    """
    origin = pd.read_excel(paths[0], origin_sheet)
    technology = pd.read_excel(paths[1], technology_sheet)
    #1: optional
    municipalities_list = municipalies_filter(origin, technology)
    #2: phase 2
    dict_origin_mun_pos, dict_technology_mun_pos, length_column= create_dicts_municipaly_position(origin,
                                                                                                 technology,
                                                                                                 municipalities_list)
    #3: phase 3
    dict_with_lists_with_mun_pos_ads_origin, dict_with_lists_with_mun_pos_ads_technology = find_address_by_mun_pos(
                                                                                             origin,
                                                                                             technology,
                                                                                             dict_origin_mun_pos,
                                                                                             dict_technology_mun_pos)
    # Close excels
    del origin
    del technology
    #4: phase 4
    #TODO ROD
    positions_list = addresses_comparator(dict_with_lists_with_mun_pos_ads_origin,
                                           dict_with_lists_with_mun_pos_ads_technology)
    #5: phase 5
    update_xlsx_from_positions_list(paths[0], origin_sheet, positions_list, length_column)

def update_xlsx_from_positions_list(path, sheet, positions_list, length_column):
    """
    A partir de esa lista, se deberá actualizar en el excel "origen" las posiciones de la lista creando una nueva
     columna. Contendrá la tecnología filtrada del excel "tecnología".
    :param positions_list: lista que contenga las posiciones [pos1,pos2,pos3,...] siendo cada posición la posición del 
    fichero origen que contiene una dirección (que el ALGORITMO ha comprobado que existe) que está en el archivo 
    "tecnología"
    """
    pt("STEP 5/5...")
    positions_list_to_update = ["TECNOLOGIA"]
    for i in range(length_column):
        if i in positions_list:
            positions_list_to_update.append("SI")
        else:
            positions_list_to_update.append("NO")
    try:
        excel = openpyxl.load_workbook(path)
        add_column(excel, origin_sheet, positions_list_to_update)
        new_path = save_excel(excel, path)
        pt("STEP 5/5 COMPLETED")
        pt("FILE " + str(new_path) + " CREATED CORRECTLY")
    except Exception:
        pt(traceback.print_exc())
        pt("STEP 5/5 FAILED")
        pt("FILE DOESN'T UPDATE CORRECTLY")

def save_excel(excel, path):
    first_part = path[:-5]
    second_part = path[-5:]

    count = 1
    new_path = first_part + "_" + str(count) + second_part
    while os.path.exists(new_path):
        count += 1
        new_path = first_part + "_" + str(count) + second_part
    excel.save(new_path)
    return new_path

def add_column(excel, sheet_name, column):
    ws = excel[sheet_name]
    new_column = ws.max_column + 1

    for rowy, value in enumerate(column, start=1):
        ws.cell(row=rowy, column=new_column, value=value)

def addresses_comparator(dict_with_lists_with_mun_pos_ads_origin,
                         dict_with_lists_with_mun_pos_ads_technology):
    """
    4º.  A partir de esos dos diccionarios, comparar mediante el ALGORITMO las direcciones. El ALGORITMO deberá comparar
     las direcciones y decidir qué direcciones son las mismas. Se deberá crear una lista que contenga las posiciones
     [pos1,pos2,pos3,...] siendo cada posición la posición del fichero origen que contiene una dirección 
     (que el ALGORITMO ha comprobado que existe) que está en el archivo "tecnología". Se retornará esa lista.

    :param dict_with_lists_with_mun_pos_ads_origin: 
    :param dict_with_lists_with_mun_pos_ads_technology: 
    :return: La lista de posiciones [pos1,pos2,pos3,...]
    """
    pt("STEP 4/5...")
    positions_list = []
    # TODO Create Phase 4
    pt("STEP 4/5 Complete")
    return positions_list

def phase_3(address_list_file, dict_file_mun_pos):
    dict = {}
    lists_pos_ads = []

    for key, value in dict_file_mun_pos.items():
        if value[0] == value [1]:
            list_pos_ads = [value[0], address_list_file[value[0]]]
            lists_pos_ads.append(list_pos_ads)
        else:
            for i in range(value[0],value[1]):
                list_pos_ads = [i,address_list_file[i]]
                lists_pos_ads.append(list_pos_ads)
            dict[key] = lists_pos_ads
    pt(dict)
    return dict


def find_address_by_mun_pos(origin, technology, dict_origin_mun_pos, dict_technology_mun_pos):
    """
    Por cada municipio, crear una lista con la forma [pos, address] por cada posición. Será una lista de listas. 
    Al final retornará un diccionario con la forma: {key=municipio : value= [[pos,address],[pos,address],...] por cada
    uno de los ficheros.
    :param origin: 
    :param technology: 
    :param dict_origin_mun_pos: 
    :param dict_technology_mun_pos: 
    :return: Retornará un diccionario con la forma: {key=municipio : value= [[pos,address],[pos,address],...] por cada
    uno de los ficheros.
    """
    pt("STEP 3/5...")
    address_column_origin = origin[[origin_column_address]]
    address_column_technology = technology[[technology_column_address]]
    address_list_origin = list(np.squeeze(np.asarray(address_column_origin.iloc[:])))
    address_list_technology = list(np.squeeze(np.asarray(address_column_technology.iloc[:])))

    dict_origin = phase_3(address_list_origin, dict_origin_mun_pos)
    dict_technology = phase_3(address_list_technology, dict_technology_mun_pos)
    pt("STEP 3/5 Complete")
    return dict_origin, dict_technology

def create_dicts_municipaly_position(origin, technology, municipalities_list=None):
    """
    2º. Método que retorne dos diccionarios que tengan la siguiente forma: {key="Municipio" : value [x,y]}. "x" e "y" son
    las posiciones donde empiezan (x) y acaban (y) los registros del "Municipio" en cuestión. Esto debe hacerse por 
    cada archivo. Es decir, debe retornar dos diccionarios, uno por cada archivo.
    :param municipalities_list: list optional
    :return: Retorna dos diccionarios, uno por cada archivo, con la forma comentada. Además, retorna la longitud 
    de la columna de municipios.
    """
    pt("STEP 2/5...")
    dict_origin = {}
    dict_technology = {}
    municipality_column_origin = origin[[origin_column_municipality]]
    municipality_column_technology = technology[[technology_column_municipality]]
    municipalities_list_origin = list(np.squeeze(np.asarray(municipality_column_origin.iloc[:])))
    municipalities_list_technology = list(np.squeeze(np.asarray(municipality_column_technology.iloc[:])))

    if municipalities_list is not None:
        dict_origin = phase_2(municipalities_list_origin, municipalities_list)
        dict_technology = phase_2(municipalities_list_technology, municipalities_list)
    else: # If not municipalities_list
        pass
    pt("STEP 2/5 Complete")
    return dict_origin, dict_technology, len(municipalities_list_origin)

def phase_2(municipalities_list_file, municipalities_list=None):
    dict = {}
    must_upload_municipality = True
    actual_municipality = ""
    last_municipality = ""
    range_ = len(municipalities_list_file)
    # File origin
    for i in range(range_):
        if i == 0 or municipalities_list_file[i] != actual_municipality and municipalities_list_file[i] \
                and type(municipalities_list_file[i]) != np.float:
            actual_municipality = municipalities_list_file[i]
            must_upload_municipality = True
            if municipalities_list_file[
                        i - 1] != actual_municipality and i != 0 and last_municipality in municipalities_list:
                dict[last_municipality] = [dict.get(last_municipality), i - 1]
        if municipalities_list_file[i] in municipalities_list:
            if must_upload_municipality: # Must be first because else execute error
                dict.update({municipalities_list_file[i]: i})
                must_upload_municipality = False
                last_municipality = actual_municipality
            if i == range_ - 1 and actual_municipality in municipalities_list:
                dict[municipalities_list_file[i]] = [dict[actual_municipality], i]


    pt(dict)
    return dict
def municipalies_filter(origin, technology):
    """
    1º. Opcional: Filtro de municipios en orden de ambos archivos. Esto es, retornar una lista con los municipios que
    # existen realmente en ambos archivos (es opcional porque no se sabe a priori si van a tener los mismos municipios).
    :param origin: excel origen
    :param technology: excel tecnología
    :return: municipalities_list = lista de municipios filtrados
    """
    pt("STEP 1/5...")
    municipality_column_origin = origin[[origin_column_municipality]]
    # This line-code get the no-repeated values in municipality_column_origin
    municipalities_list_origin = set(np.squeeze(np.asarray(municipality_column_origin.iloc[:])))
    municipality_column_technology = technology[[technology_column_municipality]]
    municipalities_list_technology = set(np.squeeze(np.asarray(municipality_column_technology.iloc[:])))
    municipalities_list = municipalities_list_origin.intersection(municipalities_list_technology)
    pt(municipalities_list)
    pt("STEP 1/5 Complete")
    return municipalities_list

algorithm(paths)